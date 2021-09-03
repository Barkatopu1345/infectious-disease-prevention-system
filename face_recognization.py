import cv2
import numpy as np
import face_recognition
import os
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
import pandas as pd
from pathlib import Path




def timeInterval(s1,s2):

    FMT = '%Y-%m-%d %H:%M:%S.%f'
    
    tdelta = datetime.strptime(str(s1), FMT) - datetime.strptime(str(s2), FMT)
    if tdelta.seconds>((3000/1000)%60):
        return True
    else: return False

def findEncodings(images): 
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList


def createSheet():
    path = ('name.xlsx')

    wb_obj = openpyxl.load_workbook(path)
    now = datetime.now()
    name = str(now.strftime('%dth%b, %Y'))


    ws = wb_obj.worksheets[-1]
    sh = str(ws).split('"')
    s = sh[1]
    if(name != s):
        wb_obj.create_sheet(name)
    else: name = s



    wb_obj.save('name.xlsx') 
    
    ws = wb_obj.worksheets[-1]
    sh = str(ws).split('"')
    s = sh[1]
    wb_obj.close()
    return s




def attendence(name,result,sheetName):
    path = ('name.xlsx')

    wb_obj = openpyxl.load_workbook(path)
    # wb_obj.iso_dates = True

    sheet = wb_obj[sheetName]


    __row__ = sheet.max_row
    column = sheet.max_column
    list = []
    
    for i in range(2,__row__+1):
        list.append(sheet.cell(i,1).value)
    

    row = __row__
    now = datetime.now()
    dtString = now.strftime("%H-%M-%S")
    sheet.cell(1,1,value = "Name")
    sheet.cell(1,1).font = Font(bold = True)

    if name not in list:
        sheet.cell(row+1,1,value=name)
        sheet.cell(row+1,2,value = "Time")
        sheet.cell(row+1,2).font = Font(bold = True)
        sheet.cell(row+1,3,value=now) 
        sheet.cell(row+1,3).number_format ="HH:MM:SS"

        sheet.cell(row+2,2,value="Matching(%)")
        sheet.cell(row+2,2).font = Font(bold = True)
        sheet.cell(row+2,3,value=result)
          

   
    elif name in list:
        timeList = []
        index = list.index(name)
        for i in range(1,column+1):
            if ((sheet.cell(index+2,i).value != None)):
                timeList.append(sheet.cell(index+2,i).value)
        
        length = len(timeList)

        oldTime = timeList[length-1]
        if timeInterval(now,oldTime) == True:
            sheet.cell(index+2, length+1, value=now)
            sheet.cell(index+2,length+1).number_format ="HH:MM:SS"
            sheet.cell(index+3, length+1, value=result)  
            

         
    wb_obj.save('name.xlsx') 
    wb_obj.close()



        
def main_face(): 
     
    path = 'image'
    images = []
    classNames = []
    myList = os.listdir(path)
    for cl in myList:
        curImg = cv2.imread(f'{path}/{cl}')
        images.append(curImg)
        classNames.append(os.path.splitext(cl)[0])

    now = datetime.now()
    dtString = now.strftime("%H:%M:%S")
    print("encoding starting time " + dtString)



    encodeListKnown = findEncodings(images)
    now = datetime.now()
    dtString = now.strftime("%H-%M-%S")
    print('Encoding Complete' + dtString)
    sheetName = createSheet()
    cap = cv2.VideoCapture(0)

    while True:
        success, img = cap.read()
        imgS = cv2.resize(img,(0,0),None,0.25,0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
        
        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS,facesCurFrame)
        for encodeFace,faceLoc in zip(encodesCurFrame,facesCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown,encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown,encodeFace)
            matchIndex = np.argmin(faceDis)

            if(faceDis[matchIndex]<0.50):
                result = faceDis[matchIndex]*100
                result = 100 - float(result)

                if matches[matchIndex]:
                    name = classNames[matchIndex].upper()
                    show = 'Matching = ' + str(int(result)) + '%  ' +name 
                    y1,x2,y2,x1 = faceLoc
                    y1, x2, y2, x1 = y1*4,x2*4,y2*4,x1*4
                    cv2.rectangle(img,(x1,y1),(x2,y2),(0,255,0),2)
                    cv2.rectangle(img,(x1,y2-35),(x2,y2),(0,255,0),cv2.FILLED)
                    cv2.putText(img, show,(x1+6,y2-6),cv2.FONT_HERSHEY_COMPLEX,0.5,(255,255,255),2)

                    attendence(name,result,sheetName)
                    
                    
        cv2.imshow('Webcam',img)
        if cv2.waitKey(2) == 27:
            break

